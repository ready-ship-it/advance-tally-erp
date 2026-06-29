"""PDF and XLS export services."""
import io
from io import BytesIO
from datetime import datetime
from flask import send_file
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def gst_export_csv(vouchers):
    """Generate a simple CSV for GSTR-1/3B data."""
    output = io.StringIO()
    output.write("Date,Voucher No,Party,GSTIN,Taxable Value,CGST,SGST,IGST,Total\n")
    for v in vouchers:
        party_name = v.party.name if v.party else "Cash"
        gstin = v.party.gstin if v.party and v.party.gstin else ""
        output.write(f"{v.voucher_date},{v.voucher_no},{party_name},{gstin},"
                     f"{v.taxable_value},{v.cgst_amount},{v.sgst_amount},{v.igst_amount},{v.grand_total}\n")
    return output.getvalue()


def gst_summary_to_xlsx(rows, fy):
    """Generate GST Monthly Summary as XLSX."""
    wb = Workbook()
    ws = wb.active
    ws.title = "GST Summary"
    
    # Header
    ws['A1'] = f"GST Monthly Summary - FY {fy}-{fy+1}"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Column headers
    headers = ["Period", "Taxable Value", "CGST", "SGST", "IGST", "Total GST", "Grand Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for row_idx, row in enumerate(rows, 4):
        ws.cell(row=row_idx, column=1).value = row.get("period", "")
        ws.cell(row=row_idx, column=2).value = row.get("taxable", 0)
        ws.cell(row=row_idx, column=3).value = row.get("cgst", 0)
        ws.cell(row=row_idx, column=4).value = row.get("sgst", 0)
        ws.cell(row=row_idx, column=5).value = row.get("igst", 0)
        ws.cell(row=row_idx, column=6).value = row.get("total_gst", 0)
        ws.cell(row=row_idx, column=7).value = row.get("total", 0)
        
        # Format as numbers
        for col in range(2, 8):
            ws.cell(row=row_idx, column=col).number_format = '#,##0.00'
    
    # Totals row
    if rows:
        total_row = len(rows) + 4
        ws.cell(row=total_row, column=1).value = "TOTAL"
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        
        for col in range(2, 8):
            cell = ws.cell(row=total_row, column=col)
            cell.value = f"=SUM({chr(64+col)}4:{chr(64+col)}{total_row-1})"
            cell.font = Font(bold=True)
            cell.number_format = '#,##0.00'
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    for col in range(2, 8):
        ws.column_dimensions[chr(64+col)].width = 18
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def gst_summary_to_pdf(rows, fy):
    """Generate GST Monthly Summary as PDF."""
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('Title', parent=styles['Normal'], fontSize=16, fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=12)
    
    elems = []
    
    # Title
    elems.append(Paragraph(f"GST Monthly Summary - FY {fy}-{fy+1}", title_style))
    elems.append(Spacer(1, 12))
    
    # Table
    table_data = [["Period", "Taxable Value", "CGST", "SGST", "IGST", "Total GST", "Grand Total"]]
    for row in rows:
        table_data.append([
            row.get("period", ""),
            f"{row.get('taxable', 0):,.2f}",
            f"{row.get('cgst', 0):,.2f}",
            f"{row.get('sgst', 0):,.2f}",
            f"{row.get('igst', 0):,.2f}",
            f"{row.get('total_gst', 0):,.2f}",
            f"{row.get('total', 0):,.2f}",
        ])
    
    # Totals
    if rows:
        totals = {
            k: sum(r[k] for r in rows)
            for k in ("taxable", "cgst", "sgst", "igst", "total_gst", "total")
        }
        table_data.append([
            "TOTAL",
            f"{totals['taxable']:,.2f}",
            f"{totals['cgst']:,.2f}",
            f"{totals['sgst']:,.2f}",
            f"{totals['igst']:,.2f}",
            f"{totals['total_gst']:,.2f}",
            f"{totals['total']:,.2f}",
        ])
    
    t = Table(table_data, colWidths=[80, 80, 70, 70, 70, 80, 80])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
    ]))
    elems.append(t)
    
    doc.build(elems)
    buf.seek(0)
    return buf


def invoice_to_pdf(voucher, settings, amt_words):
    """Generate a professional GST Tax Invoice PDF using ReportLab."""
    buf = BytesIO()
    # Reduced margins for more space
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    
    # Custom professional styles
    company_style = ParagraphStyle('CompanyTitle', parent=styles['Normal'], fontSize=16, fontName='Helvetica-Bold', leading=20)
    address_style = ParagraphStyle('Address', parent=styles['Normal'], fontSize=9, leading=11)
    title_style = ParagraphStyle('Title', parent=styles['Normal'], fontSize=14, fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=12)
    label_style = ParagraphStyle('Label', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold')
    value_style = ParagraphStyle('Value', parent=styles['Normal'], fontSize=9)
    total_label_style = ParagraphStyle('TotalLabel', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', alignment=TA_RIGHT)
    total_value_style = ParagraphStyle('TotalValue', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', alignment=TA_RIGHT)

    elems = []
    
    # 1. Header Section: Company Info & Invoice Title
    company_name = settings.get("company_name", "Your Company Name")
    company_address = settings.get("company_address", "Address Line 1\nAddress Line 2")
    company_gstin = settings.get("company_gstin", "00XXXXXXXXXXXXX")
    company_phone = settings.get("company_phone", "")
    company_email = settings.get("company_email", "")

    # Top Row: Logo/Company on Left, Invoice Type on Right
    header_data = [
        [
            [
                Paragraph(f"<b>{company_name}</b>", company_style),
                Paragraph(company_address.replace("\n", "<br/>"), address_style),
                Paragraph(f"GSTIN: {company_gstin}", address_style),
                Paragraph(f"Contact: {company_phone} | {company_email}" if company_phone else "", address_style)
            ],
            [
                Paragraph("TAX INVOICE", title_style),
                Paragraph(f"<b>Invoice No:</b> {voucher.voucher_no}", value_style),
                Paragraph(f"<b>Date:</b> {voucher.voucher_date.strftime('%d-%b-%Y')}", value_style),
                Paragraph(f"<b>Place of Supply:</b> {voucher.party.state if voucher.party else ''}", value_style)
            ]
        ]
    ]
    header_table = Table(header_data, colWidths=[330, 200])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
    ]))
    elems.append(header_table)
    elems.append(Spacer(1, 10))
    elems.append(HRFlowable(width="100%", thickness=1, color=colors.black, spaceBefore=5, spaceAfter=10))

    # 2. Billing Section: Billed To & Shipped To
    party = voucher.party
    party_name = party.name if party else "Cash / Walk-in"
    party_address = party.address if party and party.address else ""
    party_gstin = party.gstin if party and party.gstin else "N/A"
    party_state = f"{party.state} ({party.state_code})" if party else "N/A"

    billing_data = [
        [Paragraph("<b>BILLED TO:</b>", label_style), Paragraph("<b>TRANSPORT DETAILS:</b>", label_style)],
        [
            [
                Paragraph(f"<b>{party_name}</b>", value_style),
                Paragraph(party_address.replace("\n", "<br/>"), value_style),
                Paragraph(f"GSTIN: {party_gstin}", value_style),
                Paragraph(f"State: {party_state}", value_style)
            ],
            [
                Paragraph(f"Ref/PO No: {voucher.reference or 'N/A'}", value_style),
                Paragraph(f"Payment Mode: {voucher.payment_mode.title()}", value_style),
                Paragraph(f"Trans ID: {voucher.transaction_id or 'N/A'}", value_style)
            ]
        ]
    ]
    billing_table = Table(billing_data, colWidths=[330, 200])
    billing_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    elems.append(billing_table)
    elems.append(Spacer(1, 15))

    # 3. Items Table
    item_headers = ["SN", "Description of Goods", "HSN/SAC", "Qty", "Unit", "Rate", "Taxable", "GST %", "GST Amt", "Total"]
    item_data = [item_headers]
    
    for i, item in enumerate(voucher.items, 1):
        gst_amt = item.cgst_amount + item.sgst_amount + item.igst_amount
        item_data.append([
            i,
            Paragraph(item.description or (item.product.name if item.product else ""), value_style),
            item.hsn_code or "",
            f"{item.qty:,.2f}",
            item.unit or "PCS",
            f"{item.rate:,.2f}",
            f"{item.taxable_value:,.2f}",
            f"{item.gst_rate}%",
            f"{gst_amt:,.2f}",
            f"{item.line_total:,.2f}"
        ])

    # Calculate widths for A4 (595 points total, minus 60 points margin = 535)
    col_widths = [25, 160, 45, 35, 35, 50, 60, 35, 45, 60]
    t_items = Table(item_data, colWidths=col_widths, repeatRows=1)
    t_items.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#333333')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'), # SN
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),   # Desc
        ('ALIGN', (2, 1), (2, -1), 'CENTER'), # HSN
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),  # Numbers
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elems.append(t_items)

    # 4. Summary & Totals Section
    summary_data = [
        ["", "", Paragraph("Taxable Value", total_label_style), Paragraph(f"{voucher.taxable_value:,.2f}", total_value_style)],
    ]
    
    if voucher.cgst_amount:
        summary_data.append(["", "", Paragraph("CGST", total_label_style), Paragraph(f"{voucher.cgst_amount:,.2f}", total_value_style)])
    if voucher.sgst_amount:
        summary_data.append(["", "", Paragraph("SGST", total_label_style), Paragraph(f"{voucher.sgst_amount:,.2f}", total_value_style)])
    if voucher.igst_amount:
        summary_data.append(["", "", Paragraph("IGST", total_label_style), Paragraph(f"{voucher.igst_amount:,.2f}", total_value_style)])
    
    if voucher.round_off:
        summary_data.append(["", "", Paragraph("Round Off", total_label_style), Paragraph(f"{voucher.round_off:,.2f}", total_value_style)])
        
    summary_data.append(["", "", Paragraph("Grand Total", total_label_style), Paragraph(f"₹ {voucher.grand_total:,.2f}", total_value_style)])

    # Add Amount in Words on the left side of the summary
    summary_data[0][0] = Paragraph(f"<b>Amount in Words:</b><br/>{amt_words} Only", value_style)
    
    t_summary = Table(summary_data, colWidths=[250, 100, 100, 85])
    t_summary.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('SPAN', (0, 0), (1, -1)), # Merge first two columns for words
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('LINEBELOW', (2, -1), (3, -1), 1, colors.black),
    ]))
    elems.append(Spacer(1, 10))
    elems.append(t_summary)
    
    # 5. Footer Section: Terms & Signature
    elems.append(Spacer(1, 30))
    
    footer_data = [
        [
            [
                Paragraph("<b>Terms & Conditions:</b>", label_style),
                Paragraph("1. Goods once sold will not be taken back.", address_style),
                Paragraph("2. Interest @18% p.a. will be charged if payment is not made within due date.", address_style),
                Paragraph("3. Subject to local jurisdiction.", address_style),
            ],
            [
                Paragraph(f"For <b>{company_name}</b>", ParagraphStyle('Sign', parent=styles['Normal'], alignment=TA_RIGHT, fontSize=10)),
                Spacer(1, 40),
                Paragraph("Authorized Signatory", ParagraphStyle('SignLabel', parent=styles['Normal'], alignment=TA_RIGHT, fontSize=9))
            ]
        ]
    ]
    t_footer = Table(footer_data, colWidths=[330, 200])
    t_footer.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elems.append(t_footer)
    
    # Page border (Optional but professional)
    def add_border(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.black)
        canvas.setLineWidth(1)
        canvas.rect(20, 20, A4[0]-40, A4[1]-40)
        canvas.restoreState()

    doc.build(elems, onFirstPage=add_border, onLaterPages=add_border)
    buf.seek(0)
    return buf
